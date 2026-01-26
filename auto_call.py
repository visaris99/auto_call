import customtkinter as ctk
from tkinter import filedialog, messagebox
from CTkMessagebox import CTkMessagebox  # pip install CTkMessagebox
import pandas as pd
import subprocess
import datetime
import os
import sys
import json
import threading
import re
import traceback  # 에러 추적용
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 테마 설정
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# === [핵심 수정 1] 절대 경로 및 에러 로그 함수 ===
def get_base_path():
    """ 실행 파일(exe)이 있는 폴더의 절대 경로를 반환 """
    if getattr(sys, 'frozen', False):
        # PyInstaller로 빌드된 exe 실행 시
        return os.path.dirname(sys.executable)
    else:
        # 일반 파이썬 스크립트 실행 시
        return os.path.dirname(os.path.abspath(__file__))

def get_adb_path():
    """ adb.exe의 절대 경로 반환 """
    return os.path.join(get_base_path(), "adb.exe")

def log_error(msg):
    """ 에러 발생 시 프로그램 옆에 error_log.txt 파일 생성 """
    try:
        log_file = os.path.join(get_base_path(), "error_log.txt")
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {msg}\n")
    except:
        pass  # 로그 파일 생성 실패는 무시

# 윈도우 콘솔창 숨기기 옵션 (깜빡임 방지)
STARTUP_INFO = subprocess.STARTUPINFO()
STARTUP_INFO.dwFlags |= subprocess.STARTF_USESHOWWINDOW


class Config:
    RESULT_OPTIONS = ["부재중", "단선", "거절", "관심있음", "재통화예정", "접수"]
    WINDOW_SIZE = "750x700"
    STATE_FILE = "tm_state.json"
    
    COLORS = {
        "call": "#22c55e",
        "hangup": "#ef4444",
        "save": "#3b82f6",
        "skip": "#6b7280",
        "highlight_red": "FFCCCC",
        "highlight_yellow": "FFFF00"
    }


class ColumnDetector:
    @staticmethod
    def detect_columns(df: pd.DataFrame) -> dict:
        column_scores = {col: {'이름': 0, '전화번호': 0, '기타정보': 0} for col in df.columns}
        
        for col in df.columns:
            sample_data = df[col].dropna().astype(str).head(30)
            if len(sample_data) == 0:
                continue
            
            phone_count = sum(1 for v in sample_data if ColumnDetector._is_phone_number(v.strip()))
            name_count = sum(1 for v in sample_data if ColumnDetector._is_name(v.strip()))
            other_count = len(sample_data) - phone_count - name_count
            
            total = len(sample_data)
            if total > 0:
                column_scores[col]['전화번호'] = phone_count / total * 100
                column_scores[col]['이름'] = name_count / total * 100
                column_scores[col]['기타정보'] = other_count / total * 10
        
        result = {}
        used_columns = set()
        
        for role in ['전화번호', '이름', '기타정보']:
            best_col, best_score = None, -1
            for col, scores in column_scores.items():
                if col not in used_columns and scores[role] > best_score:
                    if role in ['전화번호', '이름'] and scores[role] < 50:
                        continue
                    best_score, best_col = scores[role], col
            
            if best_col is not None:
                result[role] = best_col
                used_columns.add(best_col)
        
        return result
    
    @staticmethod
    def _is_phone_number(value: str) -> bool:
        if not value:
            return False
        cleaned = re.sub(r'[\s\-\.\(\)]', '', value)
        if not cleaned.isdigit():
            return False
        length = len(cleaned)
        if 9 <= length <= 11:
            if cleaned.startswith(('010', '011', '016', '017', '018', '019', '02',
                                   '031', '032', '033', '041', '042', '043', '044',
                                   '051', '052', '053', '054', '055', '061', '062', '063', '064')):
                return True
            if cleaned.startswith('10') and length == 10:
                return True
        if length == 8 and cleaned.startswith(('15', '16', '18')):
            return True
        return False
    
    @staticmethod
    def _is_name(value: str) -> bool:
        if not value or any(char.isdigit() for char in value):
            return False
        if len(value) > 15:
            return False
        korean_only = re.sub(r'[^가-힣]', '', value)
        if 2 <= len(korean_only) <= 5:
            return True
        alpha_only = re.sub(r'[^a-zA-Z\s]', '', value)
        if 2 <= len(alpha_only) <= 20 and 1 <= len(alpha_only.split()) <= 3:
            return True
        return False
    
    @staticmethod
    def detect_header_row(df: pd.DataFrame) -> int:
        if len(df) == 0:
            return -1
        first_row = df.iloc[0]
        for value in first_row:
            if ColumnDetector._is_phone_number(str(value)):
                return -1
        header_keywords = ['이름', '성명', '성함', 'name', '전화', '번호', 'phone', 'tel', 
                          '연락처', '휴대폰', '기타', '비고', '메모']
        for value in first_row:
            if any(kw in str(value).lower() for kw in header_keywords):
                return 0
        return -1


class StateManager:
    @staticmethod
    def save(agent_name: str, file_path: str, index: int, column_mapping: dict):
        try:
            state = {
                "agent_name": agent_name, "file_path": file_path,
                "current_index": index,
                "column_mapping": {str(k): v for k, v in column_mapping.items()},
                "saved_at": datetime.datetime.now().isoformat()
            }
            # 상태 파일도 exe 위치에 저장
            path = os.path.join(get_base_path(), Config.STATE_FILE)
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(state, f, ensure_ascii=False)
        except Exception as e:
            log_error(f"Save State Error: {e}")
    
    @staticmethod
    def load() -> dict | None:
        path = os.path.join(get_base_path(), Config.STATE_FILE)
        if not os.path.exists(path):
            return None
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            log_error(f"Load State Error: {e}")
            return None
    
    @staticmethod
    def clear():
        path = os.path.join(get_base_path(), Config.STATE_FILE)
        if os.path.exists(path):
            try:
                os.remove(path)
            except:
                pass


class PhoneUtils:
    @staticmethod
    def normalize(phone: str) -> str:
        digits = re.sub(r'\D', '', str(phone))
        if digits.startswith('10') and len(digits) >= 10:
            digits = '0' + digits
        return digits
    
    @staticmethod
    def format_display(phone: str) -> str:
        n = PhoneUtils.normalize(phone)
        if len(n) == 11:
            return f"{n[:3]}-{n[3:7]}-{n[7:]}"
        elif len(n) == 10:
            if n.startswith('02'):
                return f"{n[:2]}-{n[2:6]}-{n[6:]}"
            return f"{n[:3]}-{n[3:6]}-{n[6:]}"
        return phone


class ADBController:
    @staticmethod
    def call(phone_number: str) -> bool:
        try:
            adb_exe = get_adb_path()
            if not os.path.exists(adb_exe):
                raise FileNotFoundError(f"ADB 파일을 찾을 수 없음: {adb_exe}")

            # [수정] 리스트 형태 인자 전달 + shell=False (공백 경로 에러 방지)
            cmd = [adb_exe, "shell", "am", "start", "-a", "android.intent.action.CALL", "-d", f"tel:{phone_number}"]
            
            subprocess.run(cmd, shell=False, check=True, capture_output=True, startupinfo=STARTUP_INFO)
            return True
        except Exception as e:
            log_error(f"Call Error: {e}\n{traceback.format_exc()}")
            return False
    
    @staticmethod
    def hangup() -> bool:
        try:
            adb_exe = get_adb_path()
            cmd = [adb_exe, "shell", "input", "keyevent", "6"]
            subprocess.run(cmd, shell=False, check=True, capture_output=True, startupinfo=STARTUP_INFO)
            return True
        except Exception as e:
            log_error(f"Hangup Error: {e}")
            return False
    
    @staticmethod
    def is_connected() -> bool:
        try:
            adb_exe = get_adb_path()
            # ADB 파일이 없으면 즉시 실패 처리 (불필요한 실행 방지)
            if not os.path.exists(adb_exe):
                return False

            cmd = [adb_exe, "devices"]
            result = subprocess.run(cmd, shell=False, capture_output=True, text=True, startupinfo=STARTUP_INFO)
            lines = result.stdout.strip().split('\n')
            return len(lines) > 1 and 'device' in lines[1]
        except Exception as e:
            # 연결 확인 에러는 로그가 너무 많이 쌓일 수 있으므로 필요시 주석 해제
            # log_error(f"Connection Check Error: {e}")
            return False


class CallbackScheduler:
    def __init__(self, root):
        self.root = root
        self.scheduled = []
    
    def schedule(self, name: str, phone: str, time: datetime.datetime):
        delay = int((time - datetime.datetime.now()).total_seconds() * 1000)
        if delay > 0:
            self.root.after(delay, lambda: self._notify(name, phone))
    
    def _notify(self, name: str, phone: str):
        CTkMessagebox(title="재통화 알림", message=f"📞 재통화 시간!\n\n{name}\n{phone}", icon="info")


class TelemarketingApp:
    def __init__(self, root):
        self.root = root
        self.root.withdraw()
        
        saved = StateManager.load()
        if saved and os.path.exists(saved.get('file_path', '')):
            if messagebox.askyesno("복구", f"이전 작업을 이어하시겠습니까?\n영업자: {saved['agent_name']}\n진행: {saved['current_index']}번째"):
                self.agent_name = saved['agent_name']
                self._saved_file_path = saved['file_path']
                self._saved_index = saved['current_index']
                self._saved_mapping = saved.get('column_mapping')
            else:
                self._init_new()
        else:
            self._init_new()
        
        if not self.agent_name:
            sys.exit()
        
        self.root.deiconify()
        self.root.title(f"TM 자동화")
        self.root.geometry(Config.WINDOW_SIZE)
        self.root.resizable(False, False)
        
        self.df = None
        self.current_index = 0
        self.column_mapping = {}
        self.called_numbers = set()
        self.scheduler = CallbackScheduler(root)
        
        self._build_ui()
        self._bind_shortcuts()
        
        if hasattr(self, '_saved_file_path') and self._saved_file_path:
            self._load_file(self._saved_file_path, self._saved_index, self._saved_mapping)
        
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _init_new(self):
        dialog = ctk.CTkInputDialog(text="영업자 성함을 입력해주세요:", title="로그인")
        self.agent_name = dialog.get_input()
        self._saved_file_path = None
        self._saved_index = 0
        self._saved_mapping = None
        StateManager.clear()

    def _build_ui(self):
        # 메인 컨테이너
        self.main_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # === 상단 헤더 ===
        header = ctk.CTkFrame(self.main_frame, fg_color="#1e40af", corner_radius=10)
        header.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(
            header, text=f"📞 TM 자동화 시스템", 
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="white"
        ).pack(side="left", padx=20, pady=15)
        
        ctk.CTkLabel(
            header, text=f"👤 {self.agent_name}",
            font=ctk.CTkFont(size=14),
            text_color="#93c5fd"
        ).pack(side="left", padx=10)
        
        self.adb_label = ctk.CTkLabel(header, text="● ADB", text_color="#9ca3af", font=ctk.CTkFont(size=12))
        self.adb_label.pack(side="right", padx=20)
        self._check_adb()
        
        # === 파일 로드 섹션 ===
        file_frame = ctk.CTkFrame(self.main_frame)
        file_frame.pack(fill="x", pady=(0, 10))
        
        self.btn_load = ctk.CTkButton(
            file_frame, text="📂 엑셀 DB 불러오기",
            command=self.load_excel, width=180, height=40,
            font=ctk.CTkFont(size=14)
        )
        self.btn_load.pack(side="left", padx=10, pady=10)
        
        self.lbl_status = ctk.CTkLabel(file_frame, text="파일이 로드되지 않았습니다.", text_color="#6b7280")
        self.lbl_status.pack(side="left", padx=10)
        
        self.lbl_column = ctk.CTkLabel(self.main_frame, text="", text_color="#3b82f6", font=ctk.CTkFont(size=11))
        self.lbl_column.pack()
        
        # 진행률
        self.progress = ctk.CTkProgressBar(self.main_frame, height=8)
        self.progress.pack(fill="x", pady=10)
        self.progress.set(0)
        
        # === 고객 정보 카드 ===
        info_card = ctk.CTkFrame(self.main_frame, corner_radius=15)
        info_card.pack(fill="both", expand=True, pady=10)
        
        ctk.CTkLabel(info_card, text="고객 정보", font=ctk.CTkFont(size=12), text_color="#6b7280").pack(anchor="w", padx=20, pady=(15, 5))
        
        self.lbl_index = ctk.CTkLabel(info_card, text="[0/0]", text_color="#9ca3af")
        self.lbl_index.pack(anchor="e", padx=20)
        
        self.lbl_name = ctk.CTkLabel(info_card, text="성함: -", font=ctk.CTkFont(size=24, weight="bold"))
        self.lbl_name.pack(anchor="w", padx=20, pady=(10, 5))
        
        self.lbl_phone = ctk.CTkLabel(info_card, text="📱 -", font=ctk.CTkFont(size=18), text_color="#1e40af")
        self.lbl_phone.pack(anchor="w", padx=20, pady=5)
        
        self.lbl_etc = ctk.CTkLabel(info_card, text="기타: -", text_color="#6b7280", wraplength=600, justify="left")
        self.lbl_etc.pack(anchor="w", padx=20, pady=5)
        
        self.lbl_dup = ctk.CTkLabel(info_card, text="", text_color="#ef4444", font=ctk.CTkFont(size=12, weight="bold"))
        self.lbl_dup.pack(anchor="w", padx=20, pady=(5, 15))
        
        # === 컨트롤 버튼 ===
        ctrl_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        ctrl_frame.pack(fill="x", pady=15)
        
        self.btn_call = ctk.CTkButton(
            ctrl_frame, text="📞 전화 걸기\n(F1)",
            command=self.make_call, width=140, height=60,
            fg_color=Config.COLORS["call"], hover_color="#16a34a",
            font=ctk.CTkFont(size=14, weight="bold"),
            state="disabled"
        )
        self.btn_call.pack(side="left", padx=5, expand=True)
        
        self.btn_hangup = ctk.CTkButton(
            ctrl_frame, text="❌ 통화 종료\n(F2)",
            command=self.end_call, width=140, height=60,
            fg_color=Config.COLORS["hangup"], hover_color="#dc2626",
            font=ctk.CTkFont(size=14, weight="bold"),
            state="disabled"
        )
        self.btn_hangup.pack(side="left", padx=5, expand=True)
        
        self.btn_skip = ctk.CTkButton(
            ctrl_frame, text="⏭️ 다음 고객\n(F4)",
            command=self.next_customer, width=140, height=60,
            fg_color=Config.COLORS["skip"], hover_color="#4b5563",
            font=ctk.CTkFont(size=14, weight="bold"),
            state="disabled"
        )
        self.btn_skip.pack(side="left", padx=5, expand=True)
        
        # === 결과 기록 ===
        result_card = ctk.CTkFrame(self.main_frame, corner_radius=15)
        result_card.pack(fill="x", pady=10)
        
        ctk.CTkLabel(result_card, text="상담 결과 기록", font=ctk.CTkFont(size=12), text_color="#6b7280").pack(anchor="w", padx=20, pady=(15, 10))
        
        input_row = ctk.CTkFrame(result_card, fg_color="transparent")
        input_row.pack(fill="x", padx=20, pady=(0, 15))
        
        ctk.CTkLabel(input_row, text="결과:").pack(side="left")
        self.result_var = ctk.StringVar(value=Config.RESULT_OPTIONS[0])
        self.combo = ctk.CTkComboBox(
            input_row, values=Config.RESULT_OPTIONS,
            variable=self.result_var, width=140, state="disabled",
            command=self._on_result_change
        )
        self.combo.pack(side="left", padx=(5, 20))
        
        ctk.CTkLabel(input_row, text="메모:").pack(side="left")
        self.entry_memo = ctk.CTkEntry(input_row, width=250, state="disabled", placeholder_text="메모 입력...")
        self.entry_memo.pack(side="left", padx=5)
        
        self.btn_save = ctk.CTkButton(
            input_row, text="💾 저장 (F3)",
            command=self.save_result, width=100,
            fg_color=Config.COLORS["save"], hover_color="#2563eb",
            state="disabled"
        )
        self.btn_save.pack(side="left", padx=10)
        
        # 단축키 안내
        ctk.CTkLabel(
            self.main_frame, 
            text="⌨️ 단축키: F1(전화) | F2(끊기) | F3(저장) | F4(다음)",
            text_color="#9ca3af", font=ctk.CTkFont(size=11)
        ).pack(pady=5)

    def _bind_shortcuts(self):
        self.root.bind("<F1>", lambda e: self.make_call())
        self.root.bind("<F2>", lambda e: self.end_call())
        self.root.bind("<F3>", lambda e: self.save_result())
        self.root.bind("<F4>", lambda e: self.next_customer())

    def _check_adb(self):
        def check():
            connected = ADBController.is_connected()
            self.root.after(0, lambda: self.adb_label.configure(
                text="● ADB 연결됨" if connected else "● ADB 미연결",
                text_color="#22c55e" if connected else "#ef4444"
            ))
        threading.Thread(target=check, daemon=True).start()
        self.root.after(5000, self._check_adb)

    def load_excel(self):
        try:
            path = filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")])
            if path:
                self._load_file(path)
        except Exception as e:
            log_error(f"Load Excel Error: {e}")
            CTkMessagebox(title="오류", message="파일 열기 실패 (로그 확인)", icon="cancel")

    def _load_file(self, path: str, idx: int = 0, mapping: dict = None):
        try:
            self.df = pd.read_csv(path, header=None, dtype=str) if path.endswith('.csv') else pd.read_excel(path, header=None, dtype=str)
            self.df = self.df.fillna('')
            self._current_file = path
            
            if ColumnDetector.detect_header_row(self.df) == 0:
                self.df = self.df.iloc[1:].reset_index(drop=True)
            
            self.column_mapping = {r: int(c) for r, c in mapping.items()} if mapping else ColumnDetector.detect_columns(self.df)
            
            if '전화번호' not in self.column_mapping:
                CTkMessagebox(title="오류", message="전화번호 컬럼을 찾을 수 없습니다.", icon="cancel")
                return
            
            col_names = {i: chr(65+i) for i in range(10)}
            info = [f"{r}={col_names.get(c, c)}열" for r, c in self.column_mapping.items()]
            self.lbl_column.configure(text=f"📊 자동 감지: {', '.join(info)}")
            
            self._load_called()
            self.current_index = idx
            self.lbl_status.configure(text=f"📋 {os.path.basename(path)} - {len(self.df)}명")
            self.update_display()
            
            self.btn_call.configure(state="normal")
            self.btn_skip.configure(state="normal")
        except Exception as e:
            log_error(f"File Parsing Error: {e}")
            CTkMessagebox(title="오류", message=f"파일 로드 실패: {e}", icon="cancel")

    def _load_called(self):
        filename = os.path.join(get_base_path(), f"{datetime.datetime.now():%Y-%m-%d}_{self.agent_name}_기록.xlsx")
        if os.path.exists(filename):
            try:
                wb = load_workbook(filename)
                for row in wb.active.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
                    if row[0]:
                        self.called_numbers.add(PhoneUtils.normalize(str(row[0])))
            except Exception as e:
                log_error(f"Load History Error: {e}")

    def _get_customer(self, row) -> dict:
        data = {'이름': '-', '전화번호': '', '기타정보': ''}
        for role, col in self.column_mapping.items():
            if col < len(row):
                data[role] = str(row.iloc[col])
        return data

    def update_display(self):
        if self.df is None or self.current_index >= len(self.df):
            self._show_complete()
            return
        
        try:
            c = self._get_customer(self.df.iloc[self.current_index])
            phone = PhoneUtils.normalize(c['전화번호'])
            
            self.lbl_index.configure(text=f"[{self.current_index + 1}/{len(self.df)}]")
            self.lbl_name.configure(text=c['이름'])
            self.lbl_phone.configure(text=f"📱 {PhoneUtils.format_display(c['전화번호'])}")
            self.lbl_etc.configure(text=f"기타: {c['기타정보']}" if c['기타정보'] else "")
            self.lbl_dup.configure(text="⚠️ 오늘 이미 통화한 번호" if phone in self.called_numbers else "")
            
            self.progress.set(self.current_index / len(self.df))
            self._disable_input()
            self.btn_hangup.configure(state="disabled")
            
            if hasattr(self, '_current_file'):
                StateManager.save(self.agent_name, self._current_file, self.current_index, self.column_mapping)
        except Exception as e:
            log_error(f"Display Error: {e}")

    def _show_complete(self):
        self.lbl_name.configure(text="✅ 업무 완료")
        self.lbl_phone.configure(text="새 파일을 불러와주세요")
        self.lbl_etc.configure(text="")
        self.lbl_dup.configure(text="")
        self.lbl_index.configure(text="")
        self.progress.set(1)
        self.btn_call.configure(state="disabled")
        self.btn_hangup.configure(state="disabled")
        self.btn_skip.configure(state="disabled")
        self._disable_input()
        StateManager.clear()

    def make_call(self):
        if str(self.btn_call.cget("state")) == "disabled":
            return
        c = self._get_customer(self.df.iloc[self.current_index])
        phone = PhoneUtils.normalize(c['전화번호'])
        if not phone:
            CTkMessagebox(title="경고", message="전화번호가 없습니다", icon="warning")
            return
        if ADBController.call(phone):
            self._enable_input()
            self.btn_hangup.configure(state="normal")
            self.called_numbers.add(phone)
        else:
            # 실패 시 에러 로그 파일 위치 알려줌
            log_path = os.path.join(get_base_path(), "error_log.txt")
            CTkMessagebox(title="오류", message=f"전화 실패. 로그를 확인하세요.\n{log_path}", icon="cancel")

    def end_call(self):
        if str(self.btn_hangup.cget("state")) == "disabled":
            return
        ADBController.hangup()

    def _on_result_change(self, value):
        if value == "재통화예정":
            dialog = ctk.CTkInputDialog(text="재통화 시간 (예: 14:30):", title="예약")
            time_str = dialog.get_input()
            if time_str:
                try:
                    h, m = map(int, time_str.split(':'))
                    t = datetime.datetime.now().replace(hour=h, minute=m, second=0)
                    if t < datetime.datetime.now():
                        t += datetime.timedelta(days=1)
                    c = self._get_customer(self.df.iloc[self.current_index])
                    self.scheduler.schedule(c['이름'], c['전화번호'], t)
                    self.entry_memo.delete(0, 'end')
                    self.entry_memo.insert(0, f"재통화: {time_str}")
                except:
                    CTkMessagebox(title="오류", message="시간 형식 오류", icon="warning")

    def save_result(self):
        if str(self.btn_save.cget("state")) == "disabled":
            return
        
        try:
            c = self._get_customer(self.df.iloc[self.current_index])
            result, memo = self.result_var.get(), self.entry_memo.get()
            filename = os.path.join(get_base_path(), f"{datetime.datetime.now():%Y-%m-%d}_{self.agent_name}_기록.xlsx")
            
            if not os.path.exists(filename):
                wb = Workbook()
                ws = wb.active
                ws.title = "상담기록"
                ws.append(["날짜", "이름", "전화번호", "결과", "메모"])
            else:
                wb = load_workbook(filename)
                ws = wb.active
            
            ws.append([datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), c['이름'], c['전화번호'], result, memo])
            
            # 스타일링
            row = ws.max_row
            fill = None
            if "결번" in memo:
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            elif result == "접수":
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            if fill:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = fill
            
            self._update_stats(ws, filename)
            
            wb.save(filename)
            self.next_customer()
        except PermissionError:
            CTkMessagebox(title="오류", message="엑셀 파일을 닫아주세요", icon="cancel")
            return
        except Exception as e:
            log_error(f"Save Result Error: {e}")
            CTkMessagebox(title="오류", message="저장 중 에러 발생", icon="cancel")

    def _update_stats(self, ws, filename):
        try:
            results = [r[0] for r in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True) if r[0]]
            stats = {k: results.count(k) for k in Config.RESULT_OPTIONS}
            
            border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
            align = Alignment(horizontal='center', vertical='center')
            
            ws["H2"] = os.path.basename(filename).replace(".xlsx", "").replace("_기록", "")
            ws.merge_cells("H2:I2")
            ws["H2"].alignment = align
            ws["H2"].font = Font(bold=True)
            ws["H2"].border = border
            
            for i, opt in enumerate(Config.RESULT_OPTIONS):
                for col, val in [(8, opt), (9, stats[opt])]:
                    cell = ws.cell(row=3+i, column=col, value=val)
                    cell.border = border
                    cell.alignment = align
            
            for col, val in [(8, "총통화량"), (9, len(results))]:
                cell = ws.cell(row=3+len(Config.RESULT_OPTIONS), column=col, value=val)
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = align
        except Exception as e:
            log_error(f"Stats Update Error: {e}")

    def next_customer(self):
        if str(self.btn_skip.cget("state")) == "disabled":
            return
        self.current_index += 1
        self.entry_memo.delete(0, 'end')
        self.result_var.set(Config.RESULT_OPTIONS[0])
        
        if self.df is not None and self.current_index >= len(self.df):
            self.update_display()
            CTkMessagebox(title="완료", message="모든 고객 통화 완료!", icon="check")
        else:
            self.update_display()

    def _enable_input(self):
        self.combo.configure(state="normal")
        self.entry_memo.configure(state="normal")
        self.btn_save.configure(state="normal")
        self.entry_memo.focus()

    def _disable_input(self):
        self.combo.configure(state="disabled")
        self.entry_memo.configure(state="disabled")
        self.btn_save.configure(state="disabled")

    def _on_close(self):
        if self.df is not None and self.current_index < len(self.df):
            if messagebox.askyesno("저장", "진행 상황을 저장하시겠습니까?"):
                StateManager.save(self.agent_name, self._current_file, self.current_index, self.column_mapping)
            else:
                StateManager.clear()
        self.root.destroy()


if __name__ == "__main__":
    try:
        app = ctk.CTk()
        TelemarketingApp(app)
        app.mainloop()
    except Exception as e:
        # 프로그램 시작 실패 시에도 로그 남김
        log_error(f"Fatal Startup Error: {e}\n{traceback.format_exc()}")
        messagebox.showerror("치명적 오류", f"프로그램 시작 실패\n{e}")